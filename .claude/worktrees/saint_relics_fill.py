"""Rewrite the 'Game Effect (Inventory)' column to match engine conventions.

Conventions matched:
- Fixed damage integers (no dice). Scale: fireball = 35 fire; small per-tick
  DOTs ~3-8; passive on-hit riders ~5-8; strong active ~20-30.
- No per-encounter cooldowns. Use cooldown seconds: short (30-60s),
  medium (120-300s), long-rest equivalent (86400s = 1/day).
- No reaction-roll mechanics. Trait-keyed effects use conditional_modifier
  or 'creatures with trait X are non-hostile by default'.
- HP buffs via +constitution (max_hp = 6 + con/10). Keep stat mods in the
  +5..+20 range to match existing 'blessed' (+10 will).
- Saves use existing save_stat field on conditions (will/con/dex/cha).
- Trait names lifted from the codebase: magical, occult, undead, devil,
  fey, beast, evil, cursed, noble, serpent.
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy

import os
SRC = r'C:\Users\Nolan\Downloads\DnD Tools\Saints and Relics.xlsx'
PATH = SRC  # try in-place first; fall back below if locked

EFFECTS = {
    "01": (  # Cold Flame of St. Aurelius
        "PASSIVE: in-inventory condition 'cold_flame_ward' — cold resistance "
        "(immunities {\"cold\": 0.5}); on melee hit received, attacker takes "
        "6 cold damage + 'chilled' 5s (triggered_effect on_melee_hit_received, "
        "internal cooldown 1s). Bearer visibly glows pale blue (flavor)."
    ),
    "07": (  # Ant Farm of St. Basileios
        "ACTIVE (cooldown 86400s): new ability 'Call the Faithful Swarm' — "
        "spawns a giant-ant swarm that follows the bearer for 60s, dealing "
        "4 piercing/sec to adjacent enemies and applying 'slowed' on hit. "
        "PASSIVE: condition with conditional_modifier — allies within 60px "
        "gain +5 will against effects with the 'fear' trait."
    ),
    "13": (  # Scarlet Unicorn of St. Chrysander
        "COMPANION (small system extension — persistent follower slot): a "
        "spectral red unicorn joins the party while the figurine is held. "
        "Charge attack 25 radiant on its first strike of combat; on hit, "
        "targets with 'undead' or 'devil' trait must make a will save or "
        "'frightened' 10s. PASSIVE: bearer emanates a holy aura — creatures "
        "with 'undead'/'devil' trait within 60px take 3 radiant/sec."
    ),
    "17": (  # Bed Pan of St. Caelius
        "PASSIVE: condition 'humble_servitor' grants ability 'speak_with_"
        "beetles'; creatures with 'beetle' trait will not attack the bearer "
        "first unless attacked. ACTIVE (cooldown 86400s): grants shrine spell "
        "'insect_plague' — 80px radius biting cloud for 30s; 8 piercing/sec; "
        "targets save (con) per tick or 'sickened' 6s."
    ),
    "19": (  # Harp of St. Dorian
        "ACTIVE (cooldown 86400s): grants shrine spell 'charm_serpents' — all "
        "creatures with 'serpent' trait within 240px save (will) or 'charmed' "
        "60s. PASSIVE: condition — immune to poison damage from 'serpent'-"
        "trait sources; +10 charisma vs targets with 'serpent' or 'longhorn' "
        "trait (conditional_modifier)."
    ),
    "25": (  # Minora of Eutropius
        "PASSIVE: condition 'vigil_light' — bearer is permanently illuminated "
        "and reveals 'invisible'-trait creatures within 120px (cancels their "
        "invisibility while in range). Cannot benefit from 'stealth' trait. "
        "Light cannot be extinguished by water or wind — useful flavor in "
        "flooded crypts."
    ),
    "27": (  # Fountain of Wound Wash (Eulalia)
        "PASSIVE: condition 'balm_aura' — allies within 60px heal 1 HP every "
        "6s (triggered_effect heal interval 6.0). ACTIVE (cooldown 86400s): "
        "'Anoint the Wound' — strip all conditions with 'bleeding' or "
        "'poisoned' trait from one ally and heal 15 HP."
    ),
    "29": (  # Whip of St. Phaestus
        "PASSIVE: condition — bearer's melee attacks deal +6 radiant damage "
        "vs targets with 'magical', 'occult', 'undead', or 'devil' trait "
        "(conditional_modifier on target traits). ACTIVE (cooldown 60s): "
        "'Penitent's Lash' — single 60px-range strike; on hit, suppresses "
        "all of target's abilities with 'spell' trait for 10s ('silenced')."
    ),
    "31": (  # Chalice of St. Fortunius
        "PASSIVE: condition 'guided' — immune to 'confused' and any condition "
        "with 'disorient' trait. ACTIVE (cooldown 86400s): 'Light the Way' — "
        "spawns a small yellow flame above the bearer for 3600s; creatures "
        "with 'undead' trait within 120px take 2 radiant/sec while it burns."
    ),
    "33": (  # Reed Fluter of Galenus
        "ACTIVE (cooldown 86400s): 'Sermon to the Frogs' — summons a frog "
        "swarm companion for 600s that follows the bearer; tongue attack "
        "pulls a target 30px and deals 4 bludgeoning. PASSIVE: condition "
        "grants ability 'speak_with_frogs'."
    ),
    "36": (  # Persuasive Essay of Gennadios
        "ACTIVE (cooldown 86400s): 'Recite the Essay' — one humanoid target "
        "within 60px saves (will) or 'charmed' 3600s (treats bearer as "
        "trusted advisor). PASSIVE: condition — +10 charisma; an extra +5 "
        "charisma vs targets with 'noble' or 'ruler' trait "
        "(conditional_modifier)."
    ),
    "37": (  # Cherrywood Pipe of Germanus
        "PASSIVE: condition 'hearth_warmth' — immune to 'exhausted' and "
        "'chilled'; +20 constitution (= +2 max HP); allies within 60px get "
        "cold resistance (immunities {\"cold\": 0.5}). ACTIVE (cooldown "
        "86400s): smoke the pipe — every party member within 120px heals "
        "10 HP and gains 'regenerating' 60s."
    ),
    "38": (  # Goat Milk of Galatea
        "PASSIVE: condition 'goat_milk_blessing' — immune to 'poisoned', "
        "'sickened', and 'nauseated'. ACTIVE (cooldown 86400s): drink — fully "
        "restore HP and remove one condition with 'disease' trait. COMPANION "
        "(small extension): a white nanny goat follower (small; horn-butt 6 "
        "bludgeoning; can traverse climb-trait terrain)."
    ),
    "39": (  # Straight Arrow of Gallus
        "PASSIVE: condition 'arrow_true' — bearer's ranged attacks ignore "
        "'cover' trait and gain +20 to hit chance (stat_modifier on "
        "ranged_accuracy). ACTIVE (cooldown 86400s): 'Loose the Arrow' — fire "
        "the relic; auto-hit on one target up to 400px for 30 piercing; "
        "the arrow returns to inventory after 60s."
    ),
    "42": (  # Eternal Lily of Hephaestion
        "COMPANION (extension): a squirrel familiar joins as a scout slot — "
        "high perception range, fast move_speed, fragile, no combat role. "
        "PASSIVE: condition grants shrine spell 'speak_with_animals'; "
        "creatures with 'beast' trait will not attack the bearer first unless "
        "attacked."
    ),
    "46": (  # The Last Laugh of Hilaria
        "PASSIVE: condition 'last_laugh' modeled on existing 'deny_ending' — "
        "fire resistance (immunities {\"fire\": 0.5}); first time per day the "
        "bearer would drop to 0 HP, drop to 1 instead and emit a 60px-radius "
        "laugh burst (enemies pushed 30px and 'confused' 5s). Cooldown 86400s "
        "on the save-from-death effect."
    ),
    "47": (  # Flexible Spine of Hyperion
        "PASSIVE: condition 'serpentine' — +10 dexterity; ignores penalties "
        "from 'difficult_terrain' trait; can pass through narrow gaps "
        "(utility). ACTIVE (cooldown 86400s): grants shrine spell "
        "'serpent_transformation' — for 60s, replace ability bar with a "
        "'natural_bite' variant (12 piercing + 'poisoned' on hit)."
    ),
    "52": (  # Stained Sickle of Iustinian
        "ACTIVE (cooldown 86400s): 'Rally the Oppressed' — every party member "
        "without a 'noble'/'aristocrat' trait gets condition 'oppressed_"
        "rallied' (stat_modifiers +15 strength and +15 dexterity) for 60s. "
        "PASSIVE: bearer's melee attacks deal +6 damage vs targets with "
        "'noble' or 'ruler' trait (conditional_modifier)."
    ),
    "54": (  # Bottomless Inkwell of Kallistos
        "PASSIVE: condition 'eternal_ink' — can copy spell scrolls and write "
        "inscriptions anywhere; halves the resource cost of scroll-copy "
        "actions. ACTIVE (cooldown 86400s): grants shrine spell "
        "'locate_object' — name an object; the quill points toward it within "
        "~1000px for 600s."
    ),
    "56": (  # Skull of the Favorite Dove of Lysandra
        "COMPANION (extension): a spectral dove follows the bearer (small, "
        "fragile); aura — allies within 60px heal 1 HP every 6s. ACTIVE "
        "(cooldown 86400s): 'Dove's Last Song' — every unconscious or dying "
        "ally within 240px is restored to 50% max HP."
    ),
    "66": (  # 100 Teeth of St. Polydorus
        "ACTIVE (cooldown 120s): 'Hundred-Toothed Bite' — bearer's next 5 "
        "weapon attacks each deal +4 piercing and reduce target's max HP by "
        "1 per hit (decay restored on long rest). PASSIVE: utility — bearer "
        "can chew through wood/rope/bone (interaction trigger)."
    ),
    "70": (  # Harpoon Arm of Prudentius
        "ACTIVE (cooldown 60s): 'Harpoon Cast' — throw a spectral harpoon up "
        "to 200px; 20 piercing on hit and applies 'chained'; bearer chooses "
        "to pull target adjacent OR yank self to target. PASSIVE: in-"
        "inventory condition — bearer can breathe underwater; immune to "
        "'suffocating' and any 'drowning'-trait condition."
    ),
    "72": (  # Wide-Set Hips of Pulcheria
        "PASSIVE: condition 'ox_hipped' — stat_modifiers +20 constitution "
        "(= +2 max HP), +10 strength, -20 dexterity (slower move_speed via "
        "the existing dex→speed formula). ACTIVE (cooldown 86400s): 'Ox's "
        "Lowing' — every party member within 240px gains 'shielded' 600s "
        "(absorbs the next 15 damage)."
    ),
    "74": (  # Ever-Loyal Bloodhound of Quintilian
        "COMPANION (extension, persistent): the bloodhound 'Ever-Loyal' joins "
        "the party as a full member while the figurine is held. Auto-reveals "
        "hidden creatures within 120px (scent); bite 8 piercing; immune to "
        "'charmed' and 'frightened'; returns to the bearer at end of combat."
    ),
    "78": (  # Deathmask of Severian
        "PASSIVE: condition 'unblemished_mask' — bearer immune to 'charmed' "
        "and 'frightened'; any enemy that casts an ability with 'charm' or "
        "'fear' trait targeting the bearer is itself 'silenced' 5s (backlash "
        "triggered_effect). ACTIVE (cooldown 86400s): 'Don the Mask' — for "
        "3600s gain existing 'undead_resilience' (no sleep/food) and see "
        "through 'illusion'-trait conditions within 60px."
    ),
    "80": (  # Silvertongue of Silvanus
        "ACTIVE (cooldown 86400s): new ability 'Whisper of the Tongue' — "
        "target a corpse dead less than 24h; ask up to 3 yes/no questions, "
        "answered truthfully if the corpse knew the answer in life. PASSIVE: "
        "condition — +10 charisma; bearer can understand any spoken language; "
        "immune to 'silenced'."
    ),
    "85": (  # Hammer of Theodorus
        "PASSIVE: condition — bearer's melee attacks deal +5 force damage vs "
        "targets with 'magical' trait (conditional_modifier). ACTIVE "
        "(cooldown 300s): 'Strike of Detection' — for 60s, all magical "
        "objects, active spells, and creatures with 'magical'/'occult' trait "
        "within 120px are highlighted (grants 'detect_magic'); also dispels "
        "one active 'magical'-trait condition on a target within 60px."
    ),
    "87": (  # Manacles of Theophania
        "ACTIVE (cooldown 60s): 'Bind the Tongue' — manacle one creature "
        "within 30px; target gets 'silenced' AND 'grappled' for 60s (no "
        "save). PASSIVE: condition — bearer is immune to 'poisoned' and "
        "immune to 'grappled' on self."
    ),
    "89": (  # Lariat of Valerian
        "ACTIVE (cooldown 60s): 'Thrice-Hanged Noose' — lasso one target up "
        "to 200px away (no size cap); pulls target adjacent and applies "
        "'grappled' 60s. PASSIVE: condition — HP cannot drop below 1 from "
        "'suffocating' or 'drowning' sources; +20 will save vs 'death'-"
        "trait effects."
    ),
    "91": (  # Pewter Mug of St. Walburga
        "ACTIVE (cooldown 86400s): 'Walburga's Brew' — mug refills each long "
        "rest; drink to gain existing 'mighty' and 'hasted' conditions for "
        "60s, then 'confused' for 30s afterward. PASSIVE: condition — immune "
        "to 'poisoned'; suppresses penalties from any 'intoxicated'-trait "
        "condition."
    ),
    "93": (  # Bonsai Tree of Wulfreda
        "COMPANION (small extension, tiny): the sentient bonsai sits on the "
        "bearer's shoulder; passive aura — creatures with 'evil' or 'cursed' "
        "trait that enter within 120px are highlighted to the party (grants "
        "shrine spell 'detect_evil'). ACTIVE (cooldown 86400s): 'Bonsai "
        "Bloom' — strips 'poisoned' and 'diseased' conditions from all allies "
        "within 60px and heals each for 10 HP."
    ),
    "94": (  # Brass Arm of Wenceslaus
        "PASSIVE: condition 'brass_arm' — one arm is a built-in weapon (8 "
        "bludgeoning + strength_modifier); cannot be disarmed; immune to "
        "'amputation'/'maim'-trait effects. ACTIVE (cooldown 86400s): 'Reach "
        "of the Brass Arm' — arm extends to 120px; 25 bludgeoning on hit and "
        "applies 'grappled', dragging target adjacent."
    ),
    "97": (  # Calcified Eyes of Wilibrord
        "PASSIVE: condition 'calcified_sight' — within 60px, bearer sees "
        "through 'illusion', 'invisibility', and 'polymorph' (cancels them "
        "while in range); auto-flags lies told in dialogue. Tradeoff: -10 "
        "to perception checks beyond 60px (foggy long vision). ACTIVE "
        "(cooldown 86400s): grants shrine spell 'reveal_alignment' — apply "
        "'truth_seen' to one target, exposing its 'evil'/'cursed'/'hostile' "
        "traits to the bearer."
    ),
    "99": (  # Embalmed Phallus of Wulfstan
        "PASSIVE: condition 'yew_blessed' grants shrine spell 'speak_with_"
        "plants'; +10 intelligence for herbalism/poison-craft checks; immune "
        "to poison damage from 'plant'-trait sources; foraging yields 2x "
        "herb resources. ACTIVE (cooldown 86400s): 'Wulfstan's Bloom' — every "
        "ally within 120px heals 10 HP and has one debuff condition removed."
    ),
}

wb = load_workbook(PATH)
ws = wb.active

NEW_COL = 12
HEADER_ROW = 2
DATA_START = 3

# Make sure header is set
hdr = ws.cell(HEADER_ROW, NEW_COL)
if hdr.value != "Game Effect (Inventory)":
    hdr.value = "Game Effect (Inventory)"

# Overwrite each row
filled = 0
for row in range(DATA_START, ws.max_row + 1):
    d_roll = ws.cell(row, 1).value
    if d_roll is None:
        continue
    key = str(d_roll).strip().zfill(2)
    text = EFFECTS.get(key)
    if text is None:
        print(f'WARN: no effect for d% {key} at row {row}')
        continue
    cell = ws.cell(row, NEW_COL)
    cell.value = text
    ref = ws.cell(row, 10)
    if ref.font:
        cell.font = copy(ref.font)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    if ref.border:
        cell.border = copy(ref.border)
    filled += 1

ws.column_dimensions['L'].width = 65

try:
    wb.save(PATH)
    out = PATH
except PermissionError:
    # File locked (likely open in Excel). Write to a sibling.
    base, ext = os.path.splitext(SRC)
    out = base + ' (revised)' + ext
    wb.save(out)
print(f'Wrote {out}')
print(f'Filled {filled} rows')
