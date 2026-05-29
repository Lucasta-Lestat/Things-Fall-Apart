"""Add 'Game Effect' columns to Minor Magic Items and Magic Weapons sheets
in Magical_Industrial_Revolution_Tables.xlsx.

Engine conventions matched:
- Fixed damage integers, no dice. Scale: small DOTs 3-8; consumable bursts
  10-30; weapon riders 3-6.
- Cooldowns in seconds (60s, 300s, 3600s, 86400s = 1/day).
- No reaction rolls. Use save_stat (will/con/dex/cha) for saves.
- No D&D HP scale. Stat mods in +5..+20 range matching existing 'blessed'.
- Reuse existing conditions: chilled, slowed, silenced, grappled, mighty,
  hasted, confused, sickened, frightened, regenerating, shielded,
  undead_resilience, blinded, stunned, charmed, poisoned, exhausted,
  unconscious, prone, intoxicated, paralyzed, invisible.
- Companion items flagged as small system extension.
- Pure-flavor items marked FLAVOR (no mechanic), to be honest about it.
"""
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment

SRC = r'C:\Users\Nolan\Downloads\DnD Tools\Magical_Industrial_Revolution_Tables.xlsx'

MINOR_MAGIC_ITEMS = {
    # ----- Tools -----
    1: "CONSUMABLE (1 use): use to instantly refresh one ability cooldown on the bearer, OR throw to deal 8 electric damage in a 30px radius. Glows faintly while charged.",
    2: "CONSUMABLE: spawns a stationary light source for 3600s; reveals 'invisible'-trait creatures within 120px and dispels 'darkness'-trait area effects in 60px.",
    3: "ACTIVE (cooldown 86400s, duration 3600s): grants a 'smartchain' companion (small follower, no combat role; can hold doors, fetch small items, brace traps).",
    4: "ACTIVE (cooldown 86400s): all loose debris in 180px radius gathers under the banger — reveals hidden 'small_item'-trait objects (e.g. dropped keys, coins, traps).",
    5: "UTILITY: throw any 'coal'-trait projectile 90px on use (autohit, 4 bludgeoning). Mostly flavor — useful for loading furnaces in puzzles.",
    6: "PASSIVE: in-inventory grants 'minor_motive_power' — can power one small mechanism continuously (puzzles / triggering devices). No combat use.",
    7: "WEARABLE (head slot, 1 use): on the first incoming attack that would deal bludgeoning damage, halve the damage and consume the helmet (explodes noisily).",
    8: "UTILITY: can cut/sever any 'flammable', 'fabric', 'rope', or 'soft' trait material in 1 turn; takes 60s for wood, 3600s for iron. In combat, can free an ally from 'grappled' or 'webbed' as a single action.",
    9: "WEARABLE (hands slot): PASSIVE — grants ability 'remote_hands' (interact with objects up to 300px away as if adjacent; same strength as bearer).",
    10: "CONSUMABLE: one wood-trait object/wall becomes 'softened' (treat as clay) for 3600s; melee attacks deal +20 damage vs softened-wood targets while active.",
    11: "PASSIVE in-inventory: auto-flags 'counterfeit' or 'fake' trait coins/loot when picked up; shrill sound on detect. Pure utility/anti-fraud.",
    12: "CONSUMABLE: coat a target object/floor (60px area) — area gains 'slippery' for 3600s; creatures entering save (dex) or 'prone' for 5s.",
    13: "UTILITY (100 charges, no recharge): bores a hole through 'wall' / 'stone' / 'wood' trait obstacles in 60s. Environment interaction tool.",
    14: "CONSUMABLE: throws to coat a 150px cube area in a thin rubber layer for 600s; doors/windows in the area cannot be opened from either side (traps creatures).",
    15: "ACTIVE (cooldown 86400s, duration 3600s): high-pressure air jet — extinguishes 'fire'-trait area effects within 60px; pushes 'cloud'/'gas'-trait effects away; 1-in-100 chance per use to explode (8 sonic to bearer).",
    16: "CONSUMABLE: target a creature; target gains 'depigmented' condition for 3600s — appearance becomes greyscale (no combat stat effect; useful for disguise).",
    17: "CONSUMABLE: target a creature; gains 'pigmented' for 3600s — vibrant colors; -10 to any 'stealth'-trait check.",
    18: "CONSUMABLE: coat one limb (60s to apply); grants fire resistance (immunities {\"fire\": 0.5}) for 600s; also applies 'sickened' for 60s (the rash).",
    19: "WEARABLE (badge): PASSIVE 'warning_charm' — when targeted by an ability with 'spell' trait, gains +10 to the relevant save and a shrill audio cue alerts the party.",
    20: "ACTIVE (cooldown 28800s, 3 uses/day): strike a 'cow_sized' or smaller target — removes all hair-trait coverings (cancels 'furred'/'hairy' defensive traits; deals 0 damage).",

    # ----- Personal -----
    21: "CONSUMABLE (10 uses): snap fingers — produces a small flame (lights torches, ignites 'flammable' objects within reach). No combat damage.",
    22: "CONSUMABLE (10 uses per pot): paint nails (60s to apply); gain 'detect_invisible' for 3600s — see 'invisible' and 'spectral' trait creatures.",
    23: "FLAVOR: cosmetic item. While in inventory, +5 charisma vs 'urban' or 'noble' trait NPCs (conditional_modifier).",
    24: "ACTIVE (cooldown 86400s): target the hair's owner; they fall into 'sleeping' for 28800s if at all tired (fully restores HP and clears 'exhausted').",
    25: "WEARABLE (feet slot): PASSIVE 'crabshoes' — bearer can move equally fast in any direction (no backpedal speed penalty); +10 dexterity.",
    26: "CONSUMABLE: apply to bearer; +10 charisma for 3600s (purely visual illusion). Wears off all at once.",
    27: "CONSUMABLE: produces a wig (equippable head slot); wig grants +5 charisma vs 'noble' or 'urban' NPCs. Takes 7200s to repack.",
    28: "WEARABLE: PASSIVE — immune to environmental 'rain'/'weather' conditions. ACTIVE (cooldown 60s): convert an adjacent hostile 'cloud'/'spray'-trait effect into harmless mist.",
    29: "FLAVOR: speaks event name on press. +5 charisma in social contexts where that event is relevant.",
    30: "FLAVOR: cosmetic, color-shifting fabric. After 60s of wear becomes tacky — apply -5 charisma instead.",
    31: "ACTIVE (cooldown 86400s, duration 3600s): transcribes all nearby speech to paper (any language). Utility for learning languages, recording lectures.",
    32: "WEARABLE (head slot): PASSIVE — absorbs the first 5 bludgeoning damage per encounter (regenerates over 60s). Destroyed by 'fire', 'acid', or 'magical' damage.",
    33: "UTILITY: produces 1 ice cube every 3600s. Throw to deal 3 cold damage + 'chilled' 5s.",
    34: "WEARABLE consumable (one charm wearable at a time, 1 use): on the first 'projectile'-trait hit received, 1-in-6 chance to fully negate; consumed regardless.",
    35: "WEARABLE: PASSIVE — adds one narrow inventory slot (1\" wide × 20\" deep). Anything inside still has weight; liquids soak through.",
    36: "ACTIVE (cooldown 60s): target a 'fabric'-trait object; hardens to glass for 3600s. Used to make rope into rigid bridge.",
    37: "WEARABLE: PASSIVE — bonded party members (must touch bell while making eye contact) can hear bearer's ring within 600px. Party comm/coordination, no combat effect.",
    38: "CONSUMABLE: deploys a bicycle; while riding, move_speed ×1.5 for up to 86400s (or 60s in rain/fog). Disintegrates after duration.",
    39: "UTILITY: blood-bonded padlock; only opens for bonded bearer. Crowbar (or 30+ strength check) also works.",
    40: "ACTIVE (cooldown 60s, 30 uses total): cork an open container OR a 'humanoid' nostril/ear — 1 damage and applies 'silenced' for 5s.",

    # ----- Entertainment -----
    41: "ACTIVE (cooldown 60s): mimics a named animal sound; 'beast'-trait creatures within 60px save (will) or 'confused' for 5s.",
    42: "FLAVOR: keeps accurate time. Animated clockface.",
    43: "CONSUMABLE (outdoors only): summons a 240px cube of pigeons for 600s; all creatures inside are 'blinded' and the area grants 'concealment'.",
    44: "UTILITY (requires 200gp device): captures still image of target — produces an inventory item that can serve as evidence/quest proof.",
    45: "UTILITY (requires 400gp device): captures a few seconds of motion. Flavor / quest evidence.",
    46: "WEARABLE (eye slot): PASSIVE — all living creatures appear as sheep; bearer fails 'identify_creature' checks; immune to 'visual_intimidation' / 'gaze'-trait effects.",
    47: "UTILITY: realistic mechanical fish. +10 to 'fishing' checks; can lure 'beast'-trait predators for 60s.",
    48: "CONSUMABLE: emits sparks — ignites 'flammable' objects within 30px; large versions deal 5 sonic damage in a 60px cone.",
    49: "FLAVOR: smoke shapes; +5 charisma during a 'performance'-trait scene.",
    50: "ACTIVE (cooldown 86400s, duration 600s): summons a 'lively_puppet' companion that hovers near the bearer; enemies must save (will) on first sight or 'distracted' for 5s.",
    51: "FLAVOR: tin songbirds in cage. Rumored to detect magical disasters — +5 perception vs 'magical_disaster'-trait events (when/if implemented).",
    52: "FLAVOR: pays vacuous compliments. +5 charisma for 60s after using (self-confidence buff).",
    53: "CONSUMABLE: eat — 50% chance unlocks one entry from the 'rumors' table (random information drop), 25% nothing, 25% gibberish (no effect).",
    54: "COMPANION (small extension): a perpetually tiny cat follower. No combat role; +5 charisma in social settings; consumes 1 ration/day.",
    55: "FLAVOR: tiny prancing carved figure. Decorative / quest gift.",
    56: "FLAVOR: any swing of the bell plays the next note of a tune. +5 charisma during 'performance' scene.",
    57: "FLAVOR: realistic fake wound. Used to fake an injury — +10 to 'deception' checks involving feigned damage.",
    58: "CONSUMABLE: press tie — bearer disguised as a large penguin (illusion) for 600s; once-per-tie. Useful for surprise distraction.",
    59: "FLAVOR: plays multiple instruments tinnily. +5 charisma in performance, -5 charisma to all listeners within 60px (irritating).",
    60: "FLAVOR: rainbow roses, wilt in 4 days. Gift / quest item.",

    # ----- Food and Drink -----
    61: "UTILITY: compass that always points to the nearest 'cheese'-trait object/creature within ~10km.",
    62: "ACTIVE (cooldown 86400s): summons a fly/maggot swarm within 60px for 600s; 3 piercing/sec to enemies inside; targets save (con) per tick or 'sickened' for 30s.",
    63: "WEARABLE: PASSIVE — bearer is immune to 'splash', 'spray', or 'liquid_cloud'-trait area effects (deflected 90° away on contact). Doesn't affect allies.",
    64: "ACTIVE (cooldown 86400s, 1/day): strike a 'rat'-trait creature; auto-hit, instant kill, target becomes cooked food (1 ration).",
    65: "CONSUMABLE: bearer gains 'lucky' for 60s (next attack/ability auto-succeeds), then 'sickened' for 600s.",
    66: "CONSUMABLE: eat — +10 charisma for 3600s. 'Fairy'-trait or 'vegetarian'-trait NPCs treat bearer as hostile for the duration.",
    67: "FLAVOR: joke item, cannot be peeled. Rots in 5 days.",
    68: "CONSUMABLE: inhale — removes 'exhausted' and 'fatigued' conditions; gain 'hasted' for 600s.",
    69: "CONSUMABLE: eat — if bearer has 'intoxicated' trait, heal 12 HP (otherwise just food, no heal).",
    70: "UTILITY: liquids inside cannot spill; only silver implement removes them. Anti-spill flavor.",
    71: "COMPANION (small extension, tea utility): walking teapot follower; no combat role; +5 charisma in tea / social scenes.",
    72: "ACTIVE (cooldown 86400s): target an object or 'living'-trait creature within 30px — dehydrates over 60s, dealing 15 necrotic total. Drained corpse yields 1 ration of broth.",
    73: "UTILITY: contents stay at the temperature they were sealed at. Flavor.",
    74: "CONSUMABLE (10 per pack): eat — for 60s, gain ability 'breath_flame' (5 sonic damage in a 30px cone). Harmless blue flames.",
    75: "CONSUMABLE: drink — heal 2 HP; cosmetic hair-bleaching effect.",
    76: "UTILITY: self-cleaning pot. If you tap the plate with a limb inside, deals 2 bludgeoning (scrubs vigorously).",
    77: "CONSUMABLE: removes 'hungry' / 'starving' condition for 3600s; no actual nutrition.",
    78: "UTILITY: peels vegetables/fruits. Flavor.",
    79: "FLAVOR: status symbol. While displayed, +5 charisma vs 'upper_class' trait NPCs. Juice smooths fingerprints (+10 to 'forgery' checks).",
    80: "CONSUMABLE (10 uses): adds declared flavor to food. Flavor / cooking buff.",

    # ----- Dubiously Legal -----
    81: "ACTIVE (cooldown 60s): target one brick; converts to a hidden compartment (small storage slot). Bricks can be smashed open by destroying the wall.",
    82: "CONSUMABLE: coat a 30px square wall area; bearer can see through up to 60px of solid material as glass for 3600s (one-way).",
    83: "ACTIVE (cooldown 86400s, duration 3600s): 120px radius of dim darkness around bearer; bearer gains +20 to 'stealth'-trait checks while inside; doesn't dispel 'magical_light' trait sources.",
    84: "CONSUMABLE: drink — target gains 'intoxicated_severe' (-20 to all stats for 1800s). As a thrown weapon, splashes for 5 fire damage in 30px (catches on sparks).",
    85: "CURSED in-inventory: bearer gains 'fumbling' — -20 dexterity, half attack-speed. Lasts until coin is discarded.",
    86: "WEARABLE (feet slot): PASSIVE — while bearer is moving at max move_speed, can run up vertical surfaces. Grip lost if speed drops below max; 'fall' if grip lost on sand/ivy/gravel.",
    87: "CONSUMABLE: fills a 480px area with dense fog for 600s — vision reduced to 12px; area grants 'concealment'; cancels 'ranged_targeting'.",
    88: "FLAVOR: black-market item. May give +5 charisma vs specific NPC types; small charisma penalty if discovered by 'puritan' trait NPCs.",
    89: "CONSUMABLE: throw — inflates into a rubbery copy of the bearer for 600s; functions as a decoy (50% chance enemies target it instead). Deflates harmlessly.",
    90: "UTILITY: counterfeit gin measure; +10 to 'deception' checks involving bar/inn pricing. Cheat tool.",
    91: "CONSUMABLE: strike a 'rat'-trait target; spawns 30 duplicate hostile-to-enemies rats that swarm for 60s, then vanish (small follower swarm).",
    92: "CONSUMABLE: target drinks; gains 'loathing' condition vs the first creature they see — hostile reaction, attacks if able, lasts 86400s.",
    93: "WEARABLE (head slot): PASSIVE — wearer cannot see living creatures or items they carry (effectively 'blinded' vs people). Wearer doesn't know they wear it unless they eat/sneeze/spit.",
    94: "CONSUMABLE: target drinks; 'babbling' condition for 1200s — must verbalize every thought (great for interrogation; difficult to keep secrets).",
    95: "ACTIVE (cooldown 86400s): target creature within 30px who looks inside; target 'paralyzed' for 60s by mind-melting patterns.",
    96: "CONSUMABLE: detonates — 240px area takes 30 sonic damage; loud noise alerts every nearby creature (cancels 'stealth' for the encounter); bright red smoke + sparks.",
    97: "UTILITY: cheat dice. +20 to 'gambling'-trait skill checks. Pips slither like vipers when scrutinized.",
    98: "CONSUMABLE: fills 60px area with cobwebs; area appears 'undisturbed' (covers tracks; +20 to 'disguise_area'); creatures crossing gain 'slowed' 5s.",
    99: "WEARABLE (head slot): ACTIVE (cooldown 60s, duration 600s): bearer disguised as a city watchman (illusion); 'civilian'-trait NPCs treated as 'charmed' for the duration; falls apart on inspection.",
    100: "CURSED utility (used on mounts): horse's gait becomes erratic — trot normal, canter wobbly, gallop wildly off-course. Used to fix races / embarrass rivals.",
}

MAGIC_WEAPONS = {
    1: "TRAIT (on-hit): hits dealing >3 damage apply 'glowing_mark' for 3600s — target visible to all party members through walls within 240px.",
    2: "TRAIT (while drawn): appears as harmless object (stick/ribbon/grass); first attack of an encounter from concealment is a 'surprise_attack' (+15 damage, 'stunned' 5s on hit).",
    3: "UTILITY: drips ink, can be used as a pen. Flavor.",
    4: "TRAIT (on-kill of 'ensouled'-trait target): wielder gains 'mimic_voice' condition for 3600s — can speak in target's voice with their known languages.",
    5: "TRAIT (on-hit): target saves (will) or 'charmed' for 3600s — treats wielder as friend regardless of damage dealt.",
    6: "TRAIT: deals 2× damage vs targets with 'ghost', 'spirit', or 'spectral' trait. Otherwise normal weapon.",
    7: "TRAIT (while drawn): +5 cold damage on hit. After 3 turns wielded continuously, wielder takes 1 cold/turn (increasing by 1 each subsequent turn — 'frozen_fingers').",
    8: "TRAIT (on-hit): target gets 'chromed_limb' condition for 60s — -5 to next attack (encased limb is impaired); easily broken by sundering.",
    9: "TRAIT (on-hit): target saves (con) or 'stunned' for 5s (can attack OR move, not both).",
    10: "TRAIT (while drawn): wielder gains 'invisible' condition. Weapon itself remains visible.",
    11: "TRAIT (requires two-handed swing): functions as +5 armor (shield-equivalent); can deflect 'projectile'-trait attacks; can deflect 'spell'-trait attacks back like a mirror.",
    12: "TRAIT (cooldown 60s, vertical swing): deals normal weapon damage to everything in a 30px straight line ahead.",
    13: "TRAIT: stab into a non-magical 'lock' — 50% chance to burst it open. Utility.",
    14: "TRAIT (cooldown 3600s, on-hit dealing damage): +5 fire damage as flames spurt from the wound.",
    15: "TRAIT (cooldown 30s, sideways swing): wielder dashes 60px in a straight line adding melee damage on contact. On miss, flung 60px in wrong direction.",
    16: "TRAIT (on-hit): instead of damage, target is pushed 30-180px horizontally away from wielder. Takes 5 damage if collides with wall/creature.",
    17: "TRAIT (on-hit): weapon shouts the target's true name (reveals 'hidden_identity' trait; useful for true-name magic).",
    18: "TRAIT (on-kill of cow-sized or smaller target): corpse becomes 'roasted' food (1 ration).",
    19: "TRAIT: can be sheathed inside any beard (flavor/utility, unlimited capacity).",
    20: "TRAIT (3 charges, refresh 86400s, on-hit): target saves (will) or polymorphed into a random small creature for 60s — 'stunned' first turn, fully healed in new form. After 60s, reverts.",
    21: "TRAIT (on decapitation kill of humanoid): head shrinks to a leathery trophy with a string handle. Trophy/quest item.",
    22: "TRAIT (while drawn): wielder appears as a skeleton in cloak (illusion); 'skeleton'-trait creatures won't attack first.",
    23: "TRAIT (passive while drawn): air jet from tip — can blow out 'candle'/'small_flame' trait sources, flip pages, push 'lightweight' objects.",
    24: "TRAIT: spun, points north (compass). Cannot harm 'iron' or 'steel' trait targets.",
    25: "TRAIT: cuts through ice trivially. Triple damage vs 'ice'-trait creatures.",
    26: "TRAIT (on-kill of 'ensouled' target): green flame transfers to corpse, locking it in its current position (immovable) for 3600s. Strike to release.",
    27: "TRAIT (while drawn, before first attack): observers don't believe wielder intends harm — first strike from concealment gains +10 damage and counts as 'surprise_attack'.",
    28: "TRAIT (cooldown 86400s, duration 60s): wielder spins weapon over head — lifts 30px/turn or moves 30px/turn horizontally. Wielder can do nothing else; any damage taken ends the effect.",
    29: "TRAIT: command word swaps weapon with any coin within 30px (utility — trickery / loophole tool).",
    30: "TRAIT (while drawn): wielder can steer all visible 'fire'-trait area effects (up to bonfire-size) — redirect against wind in any pattern.",
    31: "TRAIT (while drawn): 'cat'-trait creatures within 240px gain 'fascinated' (act passive, slow approach the weapon). Intelligent cats become willing to barter for it.",
    32: "TRAIT (on-hit dealing >3 damage): weapon copies one 'disease'-trait condition from target; next hit applies that disease (no save) to any 'living' target.",
    33: "TRAIT (special use, no damage): insert into target's ear — removes one 'charm', 'fear', or 'mental'-trait condition from them. Cannot harm ears or brains.",
    34: "TRAIT (2 charges, refresh 86400s): excavate a grave-size hole in soil/sand/rubble in 60s. Utility.",
    35: "TRAIT (while drawn): creatures within 60px with the 'guilty' / 'criminal' / 'wanted' trait save (will) or visibly flinch/cower/confess (interrogation tool).",
    36: "TRAIT (when drawn): plays a unique musical leitmotif. Wielder cannot be 'stealthed' while drawing. Flavor + tactical drawback.",
    37: "TRAIT (on-hit dealing >3 damage): target is cleaned/polished — removes 'dirty', 'stained', 'bloodied', or 'oiled' traits.",
    38: "TRAIT: within 240px, all 'lightning'-trait attacks redirect to the weapon. Weapon absorbs 20 damage/turn; excess transfers to wielder.",
    39: "TRAIT: command word transforms between fabric and metal forms (concealment / smuggling).",
    40: "TRAIT (while drawn): wielder is immune to 'frightened' but also 'lethargic' (-10 dexterity, cannot 'dash'-trait ability).",
    41: "TRAIT (command word): weapon flies back to wielder's hand (up to 240px). Cannot deal damage during return.",
    42: "TRAIT: every 86400s produces a silver seed. Planted, grows into a new (sterile) duplicate weapon in ~6 months of game time. Utility / quest.",
    43: "TRAIT: glows red near 'noble' or 'ruler' trait creatures. +5 damage vs them.",
    44: "TRAIT (cooldown 86400s): ask weapon a yes/no question — answers with authority (no actual knowledge); listeners save (will) or believe it.",
    45: "TRAIT (3 charges, refresh 86400s): unfurl a 30px flag with any ensign — grants nearby allies 'rally' (+5 morale stat_mod) for 600s.",
    46: "TRAIT (on-hit): produces material-specific tone — reveals 'false_door', 'forgery', or 'illusion'-trait objects within 60px.",
    47: "TRAIT (on-hit): weapon transforms into shackles around target's wrists/ankles — target gains 'grappled' + 'silenced' for 60s (no save). Command word restores weapon.",
    48: "TRAIT: cuts cakes/pies/loaves into 1-30 equal segments. Flavor utility.",
    49: "TRAIT (while drawn): wielder appears wealthier — +10 charisma vs 'commoner' / 'merchant' trait NPCs.",
    50: "TRAIT: command word transforms between silver glove and weapon — glove form grants fire and acid resistance (immunities {\"fire\": 0.5, \"acid\": 0.5}).",
    51: "TRAIT (on-hit): also heals target for 5 HP. Useful for non-lethal subdual or aiding 'healing_vulnerable' (undead) targets.",
    52: "TRAIT (on-hit, same target): each successive hit on the same 'living' target deals +1 cumulative bonus damage (resets when combat ends).",
    53: "TRAIT: only deals nonlethal damage — never kills, knocks 'unconscious' instead. Comical sound effects.",
    54: "TRAIT (while drawn): wielder is 'intoxicated' but cannot become more drunk; immune to 'poisoned'; gains 'hungry'.",
    55: "TRAIT (1 charge, refresh 86400s, on-hit): converts a 30px cube of stone to sand. Triple damage vs 'stone'-trait creatures.",
    56: "TRAIT: vibrates on every hour, double-vibrates at noon/midnight. Utility.",
    57: "TRAIT (while drawn): absorbs up to a 150px cube of 'smoke', 'poison_gas', or 'cloud'-trait atmosphere. Command word releases it.",
    58: "TRAIT (while drawn): attracts 'insect'-trait creatures within 300px (may be combat-useful via swarming, or just annoying).",
    59: "TRAIT (ability): wielder may delay turn order to act last — deals +5 damage on attacks made that round.",
    60: "TRAIT (while drawn): 60px beam of raw magic from tip — 8 damage/sec to anything in path AND 4 damage/sec to wielder. Brief brushes deal 2 damage. Tarnishes silver / cracks mirrors / melts wax in 30px.",
    61: "TRAIT (while drawn): strong cedar scent within 90px. May repel certain 'beast'-trait creatures. Flavor.",
    62: "TRAIT (while drawn): no one within 150px can write or paint — attempts produce blotches. Disrupts 'scroll'-trait casting and 'rune'-trait abilities.",
    63: "TRAIT (on-kill of human-sized or smaller): corpse launched 300px straight up; lands in 2 turns dealing 12 bludgeoning on impact in 30px area.",
    64: "TRAIT (on lethal hit vs 'ensouled' target): instead of dying, target enters 'half_life' (undead, retains memories) while weapon embedded. Dislodged if target takes >3 damage in a turn.",
    65: "TRAIT (active reaction, cooldown 30s): on incoming 'spell'-trait attack, make attack roll with -10 penalty (or -2× spell-level); on success, reflect spell to caster. Cannot 'dodge' simultaneously.",
    66: "TRAIT (on-hit): target saves (will) or 'enraged' — +5 damage vs wielder, -5 vs other targets, lasts 60s.",
    67: "TRAIT (on-hit): target saves (con) or 'warty' — cosmetic; -5 charisma for 86400s.",
    68: "TRAIT (lever): deploys invisible waterproof umbrella for 2 people. Utility flavor.",
    69: "TRAIT: can be peeled for starchy fruit (2 rations). While peeled, cannot attack; regenerates in 86400s.",
    70: "TRAIT (on-hit): one of target's teeth shatters — +1 damage to 'living'-trait targets (per-hit, non-stacking).",
    71: "TRAIT (while drawn): 'mammal'-trait creatures dog-sized or smaller within 60px who see the gem save (will) or 'paralyzed' as long as gem visible.",
    72: "TRAIT (on-kill): wielder makes a free auto-attack against nearest target — friend or foe.",
    73: "TRAIT (on-kill of cow-sized or smaller): corpse cocoons; 48hr (game-time) later releases 1000 harmless butterflies (flavor / quest).",
    74: "TRAIT (command word): swap between normal weight and 200lbs — heavy form deals +10 damage but inflicts -20 dex to wielder.",
    75: "TRAIT (on-hit): 10% chance to trigger a 30px-radius 'sphere_of_annihilation' for 5s — destroys everything in radius except this weapon and wielder's boots. Requires bare-handed wielding.",
    76: "TRAIT (3 charges, refresh 86400s): extrudes 150px of vines (functions as rope).",
    77: "TRAIT (kept chained until drawn): when drawn, all 'flammable'-trait objects within 150px catch fire (including wielder if flammable).",
    78: "TRAIT (on-hit): target tastes bitter earwax (cosmetic debuff). While drawn, dissipates 'foam' / 'froth'-trait area effects within 300px.",
    79: "TRAIT (on-hit dealing >3 damage): target 'silenced' for 5s.",
    80: "TRAIT (while drawn): drips harmless yellow custard; can fling 60px to mark a target (+10 to 'track' checks on marked targets).",
    81: "TRAIT (1 charge, refresh 86400s): launches a glowing flare 600px in a straight line; hangs in air 600s as torch-bright light source.",
    82: "TRAIT (on-kill): tiny engraving of target appears on weapon. Trophy/flavor.",
    83: "TRAIT (cooldown 172800s = 48hr): when drawn, glows red → white over 3 turns, then fires blade in a straight line for 60 damage to anything in 300px line. Wielder save (dex) or 'prone'. Regrows in 48hr.",
    84: "TRAIT (on-kill): 50% chance target rises 1 turn later as a 'zombie' (half original max HP); hostile to wielder, loses 1 HP/turn.",
    85: "TRAIT (on-hit): wielder and target can communicate telepathically for 5s. Useful for interrogation / persuasion.",
    86: "TRAIT (2 charges, refresh 86400s, on-hit): both wielder and target 'paralyzed' and 'deafened' for 30s. Loud thunderclap (cancels stealth in 300px).",
    87: "TRAIT: weapon cannot be damaged by any means. Useful as armor for itself.",
    88: "TRAIT: infinite blade extends from scabbard. -10 attack accuracy (unwieldy). Retracts in 3600s or when blade damaged. Bendy after 60px extension.",
    89: "TRAIT: floats and supports 1 person. 1 charge (refresh 86400s): drag wielder through water at 90px/turn for up to 50 turns.",
    90: "TRAIT (command word): weapon glows + hums increasingly over 6 turns (~30s). At end, glow stops. Used for intimidation — targets save (will) at end or 'frightened'.",
    91: "TRAIT (2 charges, refresh 86400s, on-hit): wielder may take any amount of damage and deal half (rounded down) to target.",
    92: "COMPANION (small extension): walks behind wielder on metal legs; while held, wielder ignores 'spider_web'/'web'-trait terrain. Follows simple verbal directions within 90px.",
    93: "TRAIT (when drawn): all wielder's worn clothes and armor fly off 30px in all directions (not damaged); jewelry unaffected. Temporary armor loss (until re-equipped).",
    94: "TRAIT: vibrates if 'ambush'-trait event imminent — wielder has 50% chance to act in surprise turn. Sometimes false-positives at birthday parties.",
    95: "TRAIT (on-hit): target saves (str) or 'earthbound' for 30s — cannot fly or lift more than one walking limb off ground.",
    96: "TRAIT (while drawn): all observers (except wielder) save (will) or believe weapon is worth 1000gp+ (covet condition). Former wielders remain affected after trade.",
    97: "UTILITY: also functions as hammer, screwdriver, crowbar, and wood saw.",
    98: "TRAIT (when drawn): spells wielder's name + title in floating sparks. Can be rewritten via 1hr of shouting. Cancels 'stealth' on draw.",
    99: "TRAIT (on-hit): wielder may instantly swap this weapon with any item in target's hand. Powerful disarm.",
    100: "TRAIT (while drawn and brandished): wielder gains 'detect_lies' — auto-flags any lie in dialogue. Wielder can ONLY speak in lies/contradictions until sheathed.",
}

assert len(MINOR_MAGIC_ITEMS) == 100, f'MMI count {len(MINOR_MAGIC_ITEMS)}'
assert len(MAGIC_WEAPONS) == 100, f'MW count {len(MAGIC_WEAPONS)}'

def write_column(ws, header_row, data_start, key_col, new_col, header_text, mapping, ref_col):
    """Write a new column populated from mapping (keyed by int value of key_col)."""
    hdr = ws.cell(header_row, new_col)
    hdr.value = header_text
    ref_hdr = ws.cell(header_row, key_col)
    if ref_hdr.font:
        hdr.font = copy(ref_hdr.font)
    if ref_hdr.fill and ref_hdr.fill.fgColor:
        hdr.fill = copy(ref_hdr.fill)
    hdr.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    if ref_hdr.border:
        hdr.border = copy(ref_hdr.border)

    filled = 0
    for row in range(data_start, ws.max_row + 1):
        key_val = ws.cell(row, key_col).value
        if key_val is None:
            continue
        try:
            key = int(key_val)
        except (TypeError, ValueError):
            continue
        text = mapping.get(key)
        if text is None:
            print(f'WARN [{ws.title}]: no effect for key {key} at row {row}')
            continue
        cell = ws.cell(row, new_col)
        cell.value = text
        ref = ws.cell(row, ref_col)
        if ref.font:
            cell.font = copy(ref.font)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if ref.border:
            cell.border = copy(ref.border)
        filled += 1
    return filled


def main():
    wb = load_workbook(SRC)

    # Minor Magic Items: cols A=d100, B=Section, C=Sub d20, D=Item, E=Description
    # Add column F = Game Effect
    mmi = wb['Minor Magic Items']
    n1 = write_column(mmi, header_row=1, data_start=2, key_col=1, new_col=6,
                      header_text='Game Effect', mapping=MINOR_MAGIC_ITEMS, ref_col=5)
    mmi.column_dimensions['F'].width = 70

    # Magic Weapons: cols A=d100, B=Weapon, C=Description
    # Add column D = Game Effect
    mw = wb['Magic Weapons']
    n2 = write_column(mw, header_row=1, data_start=2, key_col=1, new_col=4,
                      header_text='Game Effect', mapping=MAGIC_WEAPONS, ref_col=3)
    mw.column_dimensions['D'].width = 70

    try:
        wb.save(SRC)
        out = SRC
    except PermissionError:
        base, ext = os.path.splitext(SRC)
        out = base + ' (revised)' + ext
        wb.save(out)

    print(f'Wrote {out}')
    print(f'Minor Magic Items filled: {n1}')
    print(f'Magic Weapons filled: {n2}')


if __name__ == '__main__':
    main()
